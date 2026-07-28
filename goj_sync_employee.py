#!/usr/bin/env python3
"""
GOJ Employee Sheet Sync — v1.0
Reads employee Google Sheets (menus, sign-in, routes) and syncs to auth_tracker.db.
Run this whenever employee updates sheets — then regenerate daily files.
"""
import json, io, os, sys, re, sqlite3
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import defaultdict

import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────
TOKEN_PATH = '/Users/mainsobhelper/.rex_google_token.json'
DB_PATH = Path.home() / "Documents" / "goj files" / "dashboard" / "auth_tracker.db"
EMPLOYEE_DIR = Path('/tmp/goj_employee')

# Google Sheets file IDs (from Drive scan)
SHEET_IDS = {
    'first_menu': '1IfBJbKleeqA329FI3WeoFQp2xqmKYRJiy_I7RC2ZBcw',
    'second_menu': '18rs4xZHmdjt78za9tsh1bse94q-9Vn-pKXcnjID3ER0',
    'sign_in': '1ko7aVBhzLMngCuWmIZuCC5eT6WwvNEUiS8Q0vF92oy8',
    'attendance': '1XQMusZ0-rPx50QDrpf92l1mgEZdHRvmnGwpB9-moSwQ',
    'calendar': '1giUlw82mlFFfMZOvcZWqBtyB5vNntKliAamQRWzV0IE',
}

DAY_MAP = {
    'M': 'Monday', 'T': 'Tuesday', 'W': 'Wednesday',
    'TH': 'Thursday', 'F': 'Friday', 'SA': 'Saturday', 'Su': 'Sunday'
}
DAY_ABBR = {'M': 'M', 'T': 'T', 'W': 'W', 'TH': 'TH', 'F': 'F', 'SA': 'SA'}


def download_sheets():
    """Download latest Google Sheets as XLSX."""
    with open(TOKEN_PATH) as f:
        creds = json.load(f)
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload

    creds_obj = Credentials(
        token=creds['access_token'], refresh_token=creds.get('refresh_token'),
        token_uri=creds.get('token_uri', 'https://oauth2.googleapis.com/token'),
        client_id=creds['client_id'], client_secret=creds['client_secret'],
        scopes=creds['scopes']
    )
    drive = build('drive', 'v3', credentials=creds_obj)
    EMPLOYEE_DIR.mkdir(parents=True, exist_ok=True)

    names = {
        'first_menu': 'First_Shift_Menu.xlsx',
        'second_menu': 'Second_Shift_Menu.xlsx',
        'sign_in': 'SIGN_IN.xlsx',
        'attendance': 'Attendance_tracking.xlsx',
        'calendar': 'Calendar_2026.xlsx',
    }

    for key, fid in SHEET_IDS.items():
        path = EMPLOYEE_DIR / names[key]
        try:
            request = drive.files().export_media(
                fileId=fid,
                mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
            with open(path, 'wb') as f:
                f.write(fh.getvalue())
            print(f'  Downloaded: {names[key]} ({len(fh.getvalue())} bytes)')
        except Exception as e:
            print(f'  FAILED: {names[key]}: {e}')

    return names


def normalize_name(name):
    """Normalize client name for matching between employee sheet and DB."""
    if not name:
        return ''
    name = name.strip()
    # Remove trailing spaces, normalize multiple spaces
    name = re.sub(r'\s+', ' ', name)
    # Title case
    parts = name.split()
    normalized = ' '.join(p.capitalize() for p in parts)
    return normalized


def build_name_index(db):
    """Build a fuzzy lookup index from clients table."""
    cur = db.cursor()
    cur.execute("SELECT client_id, name FROM clients WHERE active = 1 OR active IS NULL")
    rows = cur.fetchall()

    index = {}
    for cid, name in rows:
        # Normalized name
        norm = normalize_name(name)
        # Last name only
        last = norm.split()[-1] if norm.split() else ''
        # First name only
        first = norm.split()[0] if norm.split() else ''

        index[norm.lower()] = (cid, norm)
        if last:
            if last.lower() not in index:
                index[last.lower()] = (cid, norm)
        if first and len(first) > 2:
            if first.lower() not in index:
                index[first.lower()] = (cid, norm)

    return index


def parse_menu_sheet(filepath, shift_label, name_index, db):
    """Parse a menu sheet (First or Second shift) and sync to client_menus."""
    print(f'\n  Parsing {filepath.name} ({shift_label})...')
    wb = openpyxl.load_workbook(filepath, data_only=True)

    cur = db.cursor()
    inserts = 0
    skipped = 0

    for sheet_name in wb.sheetnames:
        # Sheet names are like "61 M", "62 T", "63 W", "522 F", "525 M" etc.
        # Parse date from sheet name: first number = day of month, letter = day key
        match = re.match(r'(\d+)\s*([MTWHFSAu]+)', sheet_name, re.IGNORECASE)
        if not match:
            continue

        day_num = int(match.group(1))
        day_letter = match.group(2).upper()
        if day_letter == 'SU':
            day_letter = 'SA'  # Sunday → use Saturday key

        # Determine month from context — assume the sheet's context
        # For now: if day > 20 and previous month, it's likely the later month
        # Simple heuristic: check if this is a known recent date
        ws = wb[sheet_name]
        if ws.max_row < 5:
            continue

        # Read the date from row 2
        date_str = str(ws.cell(row=2, column=1).value or '')
        date_match = re.search(r'(\d+)/(\d+)', date_str)
        if date_match:
            month = int(date_match.group(1))
            day = int(date_match.group(2))
            sheet_date = date(2026, month, day)
        else:
            # Fallback: parse from sheet name
            # Use the day number — assume current/next month
            today = date.today()
            sheet_date = date(today.year, today.month, day_num)
            if sheet_date < today - timedelta(days=30):
                # Probably next month
                if today.month == 12:
                    sheet_date = date(today.year + 1, 1, day_num)
                else:
                    sheet_date = date(today.year, today.month + 1, day_num)

        week_start = sheet_date - timedelta(days=sheet_date.weekday())

        print(f'    Sheet: {sheet_name} → {sheet_date} ({DAY_MAP.get(day_letter, day_letter)})')

        # Parse menu rows (start from row 5)
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if not row or not row[0]:
                continue
            name_val = str(row[0]).strip()
            if not name_val or name_val == 'None':
                continue

            client_name = normalize_name(name_val)

            # Extract menu fields (columns vary by sheet version)
            salad = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            soup = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            main = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            side = str(row[5]).strip() if len(row) > 5 and row[5] else ''

            # Skip rows with no menu data
            if not any([salad, soup, main, side]):
                continue

            # Match to DB client
            client_id = None
            db_name = client_name

            # Try exact match
            if client_name.lower() in name_index:
                client_id, db_name = name_index[client_name.lower()]
            else:
                # Try last name match
                last = client_name.split()[-1] if client_name.split() else ''
                if last.lower() in name_index:
                    client_id, db_name = name_index[last.lower()]
                else:
                    # Try first name match
                    first = client_name.split()[0] if client_name.split() else ''
                    if first.lower() in name_index:
                        client_id, db_name = name_index[first.lower()]

            if not client_id:
                skipped += 1
                continue

            # Check if this menu entry already exists
            cur.execute("""
                SELECT id FROM client_menus
                WHERE client_id = ? AND week_start = ? AND day = ?
            """, (client_id, week_start.isoformat(), day_letter))
            existing = cur.fetchone()

            if existing:
                # Update
                cur.execute("""
                    UPDATE client_menus
                    SET salad = ?, soup = ?, main = ?, side = ?,
                        source = ?, created_at = ?
                    WHERE id = ?
                """, (salad, soup, main, side, f'employee_sync_{shift_label}',
                      datetime.utcnow().isoformat(), existing[0]))
            else:
                # Insert
                cur.execute("""
                    INSERT INTO client_menus
                    (client_id, client_name, week_start, day, salad, soup, main, side,
                     confidence, source, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0.95, ?, ?)
                """, (client_id, db_name, week_start.isoformat(), day_letter,
                      salad, soup, main, side,
                      f'employee_sync_{shift_label}', datetime.utcnow().isoformat()))
            inserts += 1

    db.commit()
    print(f'    Inserted/updated: {inserts}, Skipped (no match): {skipped}')
    return inserts, skipped


def parse_signin_routes(filepath, name_index, db):
    """Parse SIGN_IN transport tabs and sync driver routes to client_route_assignments."""
    print(f'\n  Parsing {filepath.name} for driver routes...')
    wb = openpyxl.load_workbook(filepath, data_only=True)

    cur = db.cursor()
    inserts = 0

    # Find transport tabs: "M1 TR", "M2 TR", "T1 TR", etc.
    transport_tabs = [s for s in wb.sheetnames if 'TR' in s.upper()]

    for tab in transport_tabs:
        ws = wb[tab]
        # Parse day and shift from tab name
        match = re.match(r'\s*([MTWHFSu]+)(\d)\s*TR', tab, re.IGNORECASE)
        if not match:
            continue

        day_letter = match.group(1).upper()
        shift = int(match.group(2))
        if day_letter == 'SU':
            day_letter = 'SA'

        print(f'    Tab: {tab} → {DAY_MAP.get(day_letter, day_letter)} Shift {shift}')

        # Read the two driver columns (left and right halves of the sheet)
        # Left driver: columns A-D, Right driver: columns G-J (or similar)
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 4:
                continue

            # Left side
            left_name = str(row[0]).strip() if row[0] else ''
            left_addr = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            left_phone = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            left_driver = str(row[3]).strip() if len(row) > 3 and row[3] else ''

            # Right side (columns vary — typically cols 6-9)
            right_name = str(row[6]).strip() if len(row) > 6 and row[6] else ''
            right_addr = str(row[7]).strip() if len(row) > 7 and row[7] else ''
            right_phone = str(row[8]).strip() if len(row) > 8 and row[8] else ''
            right_driver = str(row[9]).strip() if len(row) > 9 and row[9] else ''

            for client_name, addr, phone, driver in [
                (left_name, left_addr, left_phone, left_driver),
                (right_name, right_addr, right_phone, right_driver)
            ]:
                if not client_name or client_name == 'None':
                    continue

                client_name = normalize_name(client_name)
                driver = driver.strip() if driver else ''

                # Match to DB
                client_id = None
                db_name = client_name
                if client_name.lower() in name_index:
                    client_id, db_name = name_index[client_name.lower()]
                else:
                    last = client_name.split()[-1] if client_name.split() else ''
                    if last.lower() in name_index:
                        client_id, db_name = name_index[last.lower()]

                if not client_id or not driver:
                    continue

                # Upsert route assignment
                cur.execute("""
                    SELECT assignment_id FROM client_route_assignments
                    WHERE client_id = ? AND day_key = ? AND shift = ? AND is_active = 1
                """, (client_id, day_letter, shift))
                existing = cur.fetchone()

                if existing:
                    cur.execute("""
                        UPDATE client_route_assignments
                        SET driver = ?, address = ?, phone = ?, created_at = ?
                        WHERE assignment_id = ?
                    """, (driver, addr, phone, datetime.utcnow().isoformat(), existing[0]))
                else:
                    cur.execute("""
                        INSERT INTO client_route_assignments
                        (client_id, tenant_id, day_key, shift, driver, route_position, address, phone, is_active, created_at)
                        VALUES (?, 1, ?, ?, ?, 0, ?, ?, 1, ?)
                    """, (client_id, day_letter, shift, driver, addr, phone, datetime.utcnow().isoformat()))
                inserts += 1

    db.commit()
    print(f'    Route assignments synced: {inserts}')
    return inserts


def main():
    print('=' * 60)
    print('GOJ Employee Sheet Sync')
    print('=' * 60)

    # 1. Download latest sheets
    print('\n[1/4] Downloading Google Sheets...')
    names = download_sheets()

    # 2. Open DB and build name index
    print('\n[2/4] Building client name index...')
    db = sqlite3.connect(str(DB_PATH))
    name_index = build_name_index(db)
    print(f'  {len(name_index)} names indexed')

    # 3. Sync menu data
    print('\n[3/4] Syncing menu data...')
    total_inserts = 0
    total_skipped = 0

    for key, fname in [('first_menu', 'First_Shift_Menu.xlsx'), ('second_menu', 'Second_Shift_Menu.xlsx')]:
        path = EMPLOYEE_DIR / fname
        if path.exists():
            shift = 'Shift 1' if 'First' in fname else 'Shift 2'
            ins, skp = parse_menu_sheet(path, shift, name_index, db)
            total_inserts += ins
            total_skipped += skp

    # 4. Sync driver routes
    print('\n[4/4] Syncing driver routes...')
    signin_path = EMPLOYEE_DIR / 'SIGN_IN.xlsx'
    if signin_path.exists():
        route_count = parse_signin_routes(signin_path, name_index, db)

    print(f'\n{"=" * 60}')
    print(f'DONE — Menus: {total_inserts} synced, {total_skipped} unmatched')
    print(f'       Routes: {route_count} assignments')
    print(f'{"=" * 60}')

    db.close()


if __name__ == '__main__':
    main()
