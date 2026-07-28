#!/usr/bin/env python3
"""
GOJ Daily File Generator — goj_generate_daily.py
==================================================
Reads GOJ_MASTER_MENU.xlsx (or directly from DB) and generates
4 daily operation files for any given date:

  1. GOJ_{date}_SIGNIN.xlsx       — Attendance sign-in sheet
  2. GOJ_{date}_FOOD_DIST.xlsx    — Food distribution list per client
  3. GOJ_{date}_KITCHEN.xlsx      — Kitchen staff: totals per menu item
  4. GOJ_{date}_DRIVERS.xlsx      — Driver manifest

Usage:
    python3 goj_generate_daily.py                  # today
    python3 goj_generate_daily.py 2026-04-14       # specific date
    python3 goj_generate_daily.py --shift 1        # shift 1 only (default: all)
    python3 goj_generate_daily.py --out ~/Desktop  # custom output folder

Files are saved to:
  ~/Documents/goj files/dashboard/daily/YYYY-MM-DD/
"""

import sqlite3
import sys
import os
import argparse
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import defaultdict, Counter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run install_ocr_deps.command first.")
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────────
REX_DIR    = Path.home() / "Desktop" / "REX"
DB_PATH   = Path.home() / "Documents" / "goj files" / "dashboard" / "auth_tracker.db"
DASHBOARD = Path.home() / "Documents" / "goj files" / "dashboard"
DAILY_DIR = DASHBOARD / "daily"

DAYS_ORDER = ["M", "T", "W", "TH", "F", "SA"]
DAY_MAP    = {0:"M", 1:"T", 2:"W", 3:"TH", 4:"F", 5:"SA", 6:"SA"}
DAY_LABELS = {"M":"Monday","T":"Tuesday","W":"Wednesday",
              "TH":"Thursday","F":"Friday","SA":"Saturday"}

# Colors
NAVY    = "1A2742";  GOLD    = "C9A84C";  WHITE   = "FFFFFF"
GRAY_L  = "F5F5F5";  GRAY_M  = "DDDDDD";  GREEN_L = "E8F5E9"
RED_L   = "FFEBEE";  GOLD_L  = "FDF5E0";  BLUE_L  = "E3F2FD"

def _thin_border(color=GRAY_M):
    s = Side(style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def _hdr(ws, row, col, value, bold=True, bg=NAVY, fg=WHITE, size=11,
         align="center", wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", bold=bold, size=size, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    cell.border    = _thin_border()
    return cell

def _cell(ws, row, col, value, bold=False, bg=WHITE, fg="000000", size=10,
          align="left", wrap=False, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", bold=bold, size=size, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    if border:
        cell.border = _thin_border()
    return cell

def _monday(d: date) -> date:
    return d - timedelta(days=d.weekday())

# ── Load data ─────────────────────────────────────────────────────────────────
def load_clients_for_day(day_code: str, shift: int = None):
    """Return list of (name, shift, phone, plan, transport) for clients scheduled."""
    try:
        col = f"day_{day_code}_actual"
        con = sqlite3.connect(str(DB_PATH))
        q = f"SELECT name, shift, phone, COALESCE(plan_canonical, plan_raw, ''), COALESCE(transportation, '') FROM clients WHERE active=1 AND {col}>0"
        if shift:
            q += f" AND shift={shift}"
        q += " ORDER BY shift, name"
        rows = con.execute(q).fetchall()
        con.close()
        return rows
    except Exception as e:
        print(f"  Warning: could not load clients: {e}")
        return []

def load_menus_for_day(day_code: str, week_start: str):
    """Return dict: client_name → {salad, soup, main, side}"""
    menus = {}
    try:
        con = sqlite3.connect(str(DB_PATH))
        rows = con.execute(
            "SELECT client_name, salad, soup, main, side FROM client_menus "
            "WHERE week_start=? AND day=?",
            (week_start, day_code)
        ).fetchall()
        con.close()
        for (name, salad, soup, main, side) in rows:
            menus[name] = {
                "salad": salad or "—",
                "soup":  soup  or "—",
                "main":  main  or "—",
                "side":  side  or "—",
            }
    except Exception as e:
        print(f"  Warning: could not load menus: {e}")
    return menus

def _title_block(ws, title, subtitle, date_str):
    ws.merge_cells("A1:J1")
    ws["A1"] = "GARDEN OF JOY ADULT DAY CARE"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:J2")
    ws["A2"] = f"{title}   |   {subtitle}   |   {date_str}"
    ws["A2"].font = Font(name="Arial", bold=True, size=12, color=WHITE)
    ws["A2"].fill = PatternFill("solid", fgColor=GOLD)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

# ═══════════════════════════════════════════════════════════════════
# 1. SIGN-IN SHEET
# ═══════════════════════════════════════════════════════════════════
def generate_signin(target_date: date, shift: int, out_dir: Path) -> Path:
    day_code  = DAY_MAP.get(target_date.weekday(), "M")
    day_label = DAY_LABELS[day_code]
    date_str  = target_date.strftime("%B %d, %Y")
    shift_str = f"Shift {shift}"
    clients   = load_clients_for_day(day_code, shift)
    total     = len(clients)
    per_page  = 11
    pages     = max(1, (total + per_page - 1) // per_page)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sign-In"

    # Set print area and page setup
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'

    # Column widths — landscape, large print
    widths = [5, 36, 32, 7, 16, 16, 36]  # No, Name, Plan, TR, Time In, Time Out, Signature
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ADDRESS_LINE = "3152 Brighton 6 St, Brooklyn NY 11235 | Garden of Joy Adult Day Care Center"

    current_row = 1
    for page in range(1, pages + 1):
        start_idx = (page - 1) * per_page
        page_clients = clients[start_idx:start_idx + per_page]

        # Title block — light gray to match official template
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        _cell(ws, current_row, 1, "GARDEN OF JOY ADULT DAY CARE CENTER — SIGN-IN SHEET",
              bold=True, size=16, bg=GRAY_M, fg="000000", align="center")
        ws.row_dimensions[current_row].height = 34
        current_row += 1

        # Date/Shift/Total/Page line
        info = f"Date: {date_str}   Shift: {shift}   Total: {total}   Page {page}/{pages}"
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        _cell(ws, current_row, 1, info, bold=False, size=11, bg=WHITE, fg="000000", align="left")
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Column headers — gray to match official template
        hdrs = ["No", "Name", "Plan", "TR", "Time In", "Time Out", "Signature"]
        for col, h in enumerate(hdrs, 1):
            _hdr(ws, current_row, col, h, size=12, bg=GRAY_M, fg="000000")
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # Client rows
        for i, row_data in enumerate(page_clients):
            name, sh, phone, plan, transport = row_data
            bg = GRAY_L if (start_idx + i) % 2 == 0 else WHITE
            tr_flag = "N/TR" if (transport or "").upper() in ("N/TR", "NTR", "NO", "NONE", "") else "TR"

            _cell(ws, current_row, 1, start_idx + i + 1, align="center", bg=bg, size=12)
            _cell(ws, current_row, 2, name, bold=True, bg=bg, size=13)
            _cell(ws, current_row, 3, plan or "", bg=bg, size=11, wrap=True)
            _cell(ws, current_row, 4, tr_flag, align="center", bg=bg, size=12)
            _cell(ws, current_row, 5, "", bg=bg)   # Time In — blank for staff
            _cell(ws, current_row, 6, "", bg=bg)   # Time Out — blank for staff
            _cell(ws, current_row, 7, "", bg=bg)   # Signature — blank
            ws.row_dimensions[current_row].height = 36
            current_row += 1

        # Address footer on every page
        _cell(ws, current_row, 1, ADDRESS_LINE, size=9, fg="888888", align="center", bg=WHITE)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Last page: totals and notes
        if page == pages:
            current_row += 1  # blank row
            _cell(ws, current_row, 1, f"Total present: ______   Staff signature: ___________________________   Date: __________",
                  bold=True, size=12, align="left", bg=WHITE)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            ws.row_dimensions[current_row].height = 26

        # Page break between pages
        if page < pages:
            current_row += 2  # gap between pages

    ws.freeze_panes = "A4"
    fname = out_dir / f"GOJ_{target_date.isoformat()}_S{shift}_SIGNIN.xlsx"
    wb.save(str(fname))
    print(f"  📋 Sign-in:      {fname.name}  ({total} clients, {pages} page(s))")
    return fname

# ═══════════════════════════════════════════════════════════════════
# 2. FOOD DISTRIBUTION SHEET
# ═══════════════════════════════════════════════════════════════════
def generate_food_dist(target_date: date, shift: int, out_dir: Path) -> Path:
    day_code   = DAY_MAP.get(target_date.weekday(), "M")
    day_label  = DAY_LABELS[day_code]
    week_start = _monday(target_date).isoformat()
    date_str   = target_date.strftime("%B %d, %Y")
    clients    = load_clients_for_day(day_code, shift)
    menus      = load_menus_for_day(day_code, week_start)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Food Distribution"

    _title_block(ws, "FOOD DISTRIBUTION SHEET",
                 f"{day_label} Shift {shift}", date_str)

    hdrs   = ["#", "Client Name", "Shift", "Salad / Салат", "Soup / Суп",
              "Main / Основное", "Side / Гарнир", "Notes", "✓ Served"]
    widths = [4, 26, 7, 22, 22, 28, 20, 18, 9]
    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws, 3, col, h, wrap=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 28

    no_menu_clients = []
    for i, row_data in enumerate(clients, 1):
        name = row_data[0]
        sh = row_data[1] if len(row_data) > 1 else 1
        menu = menus.get(name)
        bg   = GRAY_L if i % 2 == 0 else WHITE
        row  = 3 + i

        _cell(ws, row, 1, i,    align="center", bg=bg)
        _cell(ws, row, 2, name, bold=True, bg=bg)
        _cell(ws, row, 3, f"S{sh}", align="center", bg=bg)

        if menu:
            for j, course in enumerate(["salad","soup","main","side"], 4):
                val = menu.get(course, "—")
                _cell(ws, row, j, val, bg=GOLD_L if val != "—" else RED_L,
                      wrap=True, size=9)
        else:
            for j in range(4, 8):
                _cell(ws, row, j, "⚠ No menu on file", bg=RED_L,
                      fg="CC0000", size=9, align="center")
            no_menu_clients.append(name)

        _cell(ws, row, 8, "", bg=bg)   # Notes
        _cell(ws, row, 9, "", bg=bg, align="center")  # Served checkbox
        ws.row_dimensions[row].height = 26

    total_row = 3 + len(clients) + 1
    ws.merge_cells(f"A{total_row}:C{total_row}")
    ws[f"A{total_row}"] = f"TOTAL: {len(clients)} clients  |  {len(no_menu_clients)} missing menu"
    ws[f"A{total_row}"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws[f"A{total_row}"].fill = PatternFill("solid", fgColor=NAVY if not no_menu_clients else "CC0000")
    ws[f"A{total_row}"].alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A4"
    fname = out_dir / f"GOJ_{target_date.isoformat()}_S{shift}_FOOD_DIST.xlsx"
    wb.save(str(fname))
    missing = f"  ⚠ {len(no_menu_clients)} missing menu" if no_menu_clients else ""
    print(f"  🍽  Food dist:   {fname.name}  ({len(clients)} clients{missing})")
    return fname

# ═══════════════════════════════════════════════════════════════════
# 3. KITCHEN STAFF TOTALS
# ═══════════════════════════════════════════════════════════════════
def generate_kitchen(target_date: date, out_dir: Path) -> Path:
    """Kitchen sheet covers ALL shifts — kitchen preps for everyone."""
    day_code   = DAY_MAP.get(target_date.weekday(), "M")
    day_label  = DAY_LABELS[day_code]
    week_start = _monday(target_date).isoformat()
    date_str   = target_date.strftime("%B %d, %Y")

    # Load all clients for this day (both shifts)
    clients = load_clients_for_day(day_code, shift=None)
    menus   = load_menus_for_day(day_code, week_start)

    # Count totals
    totals = {"salad": Counter(), "soup": Counter(),
              "main":  Counter(), "side": Counter()}
    for row_data in clients:
        name, sh, phone = row_data[0], row_data[1], row_data[2]
        m = menus.get(name, {})
        for course in totals:
            item = m.get(course, "")
            if item and item != "—":
                totals[course][item] += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kitchen"

    _title_block(ws, "KITCHEN PRODUCTION SHEET", day_label, date_str)

    row = 4
    course_labels = {
        "salad": "SALADS / САЛАТЫ",
        "soup":  "SOUPS / СУПЫ",
        "main":  "MAINS / ОСНОВНЫЕ",
        "side":  "SIDES / ГАРНИРЫ",
    }
    course_colors = {"salad": "E8F5E9", "soup": "E3F2FD",
                     "main":  GOLD_L,   "side": "F3E5F5"}

    for course, label in course_labels.items():
        # Section header
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Arial", bold=True, size=12, color=WHITE)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=NAVY)
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center",
                                            indent=1)
        ws.row_dimensions[row].height = 22
        row += 1

        # Column headers
        for col, h in enumerate(["Menu Item", "Count", "Prepared ✓",
                                  "Served ✓", "Notes"], 1):
            _hdr(ws, row, col, h, bg=GOLD, fg=NAVY, size=10)
        ws.row_dimensions[row].height = 20
        row += 1

        items = totals[course]
        if items:
            bg_c = course_colors[course]
            for i, (item, count) in enumerate(sorted(items.items(),
                                                      key=lambda x: -x[1])):
                bg = bg_c if i % 2 == 0 else WHITE
                _cell(ws, row, 1, item,  bold=True, bg=bg, size=10)
                _cell(ws, row, 2, count, align="center", bg=bg, size=10, bold=True)
                _cell(ws, row, 3, "",    bg=bg)
                _cell(ws, row, 4, "",    bg=bg)
                _cell(ws, row, 5, "",    bg=bg)
                ws.row_dimensions[row].height = 20
                row += 1
        else:
            ws.merge_cells(f"A{row}:E{row}")
            ws[f"A{row}"] = "No menu data for this day"
            ws[f"A{row}"].font = Font(name="Arial", italic=True, size=10,
                                      color="888888")
            ws[f"A{row}"].alignment = Alignment(horizontal="center")
            row += 1

        # Total line
        total_count = sum(items.values())
        ws.merge_cells(f"A{row}:A{row}")
        _cell(ws, row, 1, f"SUBTOTAL:", bold=True, bg=GRAY_L, size=10)
        _cell(ws, row, 2, total_count, bold=True, bg=GRAY_L, align="center")
        for col in range(3, 6):
            _cell(ws, row, col, "", bg=GRAY_L)
        ws.row_dimensions[row].height = 20
        row += 2   # blank line between sections

    # Grand total
    grand_total = len(clients)
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = f"TOTAL CLIENTS TODAY (ALL SHIFTS): {grand_total}"
    ws[f"A{row}"].font = Font(name="Arial", bold=True, size=12, color=WHITE)
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=NAVY)
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center",
                                        indent=1)
    ws.row_dimensions[row].height = 22

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20

    fname = out_dir / f"GOJ_{target_date.isoformat()}_KITCHEN.xlsx"
    wb.save(str(fname))
    print(f"  🍳 Kitchen:      {fname.name}  ({grand_total} clients total)")
    return fname

# ═══════════════════════════════════════════════════════════════════
# 4. DRIVER MANIFEST
# ═══════════════════════════════════════════════════════════════════
def generate_drivers(target_date: date, shift: int, out_dir: Path) -> Path:
    day_code  = DAY_MAP.get(target_date.weekday(), "M")
    day_label = DAY_LABELS[day_code]
    date_str  = target_date.strftime("%B %d, %Y")
    clients   = load_clients_for_day(day_code, shift)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drivers"

    _title_block(ws, "DRIVER MANIFEST", f"{day_label} Shift {shift}", date_str)

    hdrs   = ["#", "Client Name", "Shift", "Phone", "Address",
              "Pickup Time", "Drop-off", "Driver", "✓ Picked Up", "Notes"]
    widths = [4, 24, 7, 14, 30, 12, 12, 16, 12, 18]
    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws, 3, col, h, wrap=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 28

    for i, row_data in enumerate(clients, 1):
        name = row_data[0]
        sh = row_data[1] if len(row_data) > 1 else 1
        phone = row_data[2] if len(row_data) > 2 else ''
        plan = row_data[3] if len(row_data) > 3 else ''
        transport = row_data[4] if len(row_data) > 4 else ""
        bg  = GRAY_L if i % 2 == 0 else WHITE
        row = 3 + i
        _cell(ws, row, 1,  i,              align="center", bg=bg)
        _cell(ws, row, 2,  name,           bold=True, bg=bg)
        _cell(ws, row, 3,  f"S{sh}",       align="center", bg=bg)
        _cell(ws, row, 4,  phone or "",    bg=bg)
        _cell(ws, row, 5,  "",             bg=bg)   # Address (from client file)
        _cell(ws, row, 6,  "",             bg=bg)   # Pickup time
        _cell(ws, row, 7,  "",             bg=bg)   # Drop-off
        _cell(ws, row, 8,  "",             bg=bg)   # Driver name
        _cell(ws, row, 9,  "",             bg=bg, align="center")  # Checkbox
        _cell(ws, row, 10, "",             bg=bg)   # Notes
        ws.row_dimensions[row].height = 22

    total_row = 3 + len(clients) + 1
    ws.merge_cells(f"A{total_row}:D{total_row}")
    ws[f"A{total_row}"] = f"TOTAL PICKUPS: {len(clients)}"
    ws[f"A{total_row}"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws[f"A{total_row}"].fill = PatternFill("solid", fgColor=NAVY)
    ws[f"A{total_row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[total_row].height = 22

    ws.freeze_panes = "A4"
    fname = out_dir / f"GOJ_{target_date.isoformat()}_S{shift}_DRIVERS.xlsx"
    wb.save(str(fname))
    print(f"  🚗 Drivers:      {fname.name}  ({len(clients)} clients)")
    return fname

# ── Entry point ───────────────────────────────────────────────────────────────
def generate_all(target_date: date = None, shift_filter: int = None,
                 out_dir: Path = None):
    if target_date is None:
        target_date = date.today()
    if out_dir is None:
        out_dir = DAILY_DIR / target_date.isoformat()
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── Drive-first: sync attendance + menus before generating anything ────
    if str(REX_DIR) not in sys.path:
        sys.path.insert(0, str(REX_DIR))
    try:
        from CC_drive_preflight import preflight
        print("🔍 Running Drive preflight to sync live data...")
        pf = preflight(target_date.isoformat())
        print(f"   Attendance: S1={pf['stats'].get('s1_attendance','?')} S2={pf['stats'].get('s2_attendance','?')}")
        print(f"   Menus: S1={pf['stats'].get('s1_menu','?')} S2={pf['stats'].get('s2_menu','?')}")
        if pf.get('no_menu'):
            print(f"   ⚠️  {len(pf['no_menu'])} clients missing menus")
    except Exception as e:
        print(f"   ⚠️  Preflight skipped (non-fatal): {e}")

    day_code = DAY_MAP.get(target_date.weekday(), "M")
    day_label = DAY_LABELS[day_code]
    print(f"\n  📅 Generating daily files for {target_date} ({day_label})")
    print(f"  📂 Output: {out_dir}")
    print()

    files = []
    shifts = [shift_filter] if shift_filter else [1, 2]

    for shift in shifts:
        files.append(generate_signin(target_date, shift, out_dir))
        files.append(generate_food_dist(target_date, shift, out_dir))
        files.append(generate_drivers(target_date, shift, out_dir))

    # Kitchen is one file for all shifts
    files.append(generate_kitchen(target_date, out_dir))

    print(f"\n  ✅ {len(files)} files generated in: {out_dir}")
    return files

def main():
    parser = argparse.ArgumentParser(description="Generate GOJ daily operation files")
    parser.add_argument("date",   nargs="?",
                        help="Date YYYY-MM-DD (default: today)")
    parser.add_argument("--shift", type=int, choices=[1,2],
                        help="Shift filter (default: generate both)")
    parser.add_argument("--out",  type=str,
                        help="Output directory (default: dashboard/daily/DATE/)")
    args = parser.parse_args()

    target_date = date.fromisoformat(args.date) if args.date else date.today()
    out_dir = Path(args.out) if args.out else None

    generate_all(target_date, args.shift, out_dir)

if __name__ == "__main__":
    main()
