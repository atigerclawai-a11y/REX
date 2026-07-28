#!/usr/bin/env python3
"""
GOJ Master Menu — goj_master_menu.py
=====================================
Reads the client_menus table from auth_tracker.db and writes
GOJ_MASTER_MENU.xlsx in the dashboard folder.

That spreadsheet is the central base for generating daily files:
  - Sign-in sheets
  - Food distribution sheets
  - Kitchen staff totals
  - Driver manifests

Usage:
    python3 goj_master_menu.py            # sync DB → master spreadsheet
    python3 goj_master_menu.py --week     # sync only current week
    python3 goj_master_menu.py --show     # print current week to terminal

Called automatically by goj_signin_intake.py when a menu file is processed.
"""

import sqlite3
import sys
import os
import argparse
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule
except ImportError:
    print("ERROR: openpyxl not installed. Run install_ocr_deps.command first.")
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────────
DB_PATH      = Path.home() / "Documents" / "goj files" / "dashboard" / "auth_tracker.db"
DASHBOARD    = Path.home() / "Documents" / "goj files" / "dashboard"
MASTER_FILE  = DASHBOARD / "GOJ_MASTER_MENU.xlsx"

DAYS_ORDER   = ["M", "T", "W", "TH", "F", "SA"]
DAY_LABELS   = {"M": "Monday", "T": "Tuesday", "W": "Wednesday",
                "TH": "Thursday", "F": "Friday", "SA": "Saturday"}
COURSES      = ["salad", "soup", "main", "side"]
COURSE_LABELS = {"salad": "Salad / Салат", "soup": "Soup / Суп",
                 "main": "Main / Основное", "side": "Side / Гарнир"}

# ── Colors (GOJ navy + gold palette) ──────────────────────────────────────────
NAVY       = "1A2742"
GOLD       = "C9A84C"
LIGHT_NAVY = "D6DCF0"
LIGHT_GOLD = "FDF5E0"
WHITE      = "FFFFFF"
GRAY_LIGHT = "F5F5F5"
GRAY_MED   = "DDDDDD"
GREEN_LIGHT = "E8F5E9"
RED_LIGHT  = "FFEBEE"

def _style(ws, cell_ref, bold=False, color=None, bg=None, align="left",
           border=False, size=11, wrap=False, italic=False):
    cell = ws[cell_ref] if isinstance(cell_ref, str) else cell_ref
    cell.font = Font(name="Arial", bold=bold, size=size,
                     color=color or "000000", italic=italic)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    if border:
        thin = Side(style="thin", color=GRAY_MED)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

def _monday(d: date) -> date:
    return d - timedelta(days=d.weekday())

def _load_clients():
    """Return dict of client_name → {shift, days: set}"""
    clients = {}
    try:
        con = sqlite3.connect(str(DB_PATH))
        rows = con.execute(
            "SELECT name, shift, day_M_actual, day_T_actual, day_W_actual, "
            "day_TH_actual, day_F_actual, day_Su_actual FROM clients WHERE active=1"
        ).fetchall()
        con.close()
        for r in rows:
            name, shift = r[0], r[1] or 1
            days = set()
            for i, code in enumerate(["M","T","W","TH","F","SA"]):
                if r[2+i] and float(r[2+i]) > 0:
                    days.add(code)
            clients[name] = {"shift": shift, "days": days}
    except Exception as e:
        print(f"  Warning: could not load clients: {e}")
    return clients

def _load_menus(week_start: str = None):
    """Return list of menu rows for the given week (or all weeks if None)."""
    try:
        con = sqlite3.connect(str(DB_PATH))
        if week_start:
            rows = con.execute(
                "SELECT client_name, week_start, day, salad, soup, main, side, confidence "
                "FROM client_menus WHERE week_start=? ORDER BY client_name, day",
                (week_start,)
            ).fetchall()
        else:
            rows = con.execute(
                "SELECT client_name, week_start, day, salad, soup, main, side, confidence "
                "FROM client_menus ORDER BY week_start DESC, client_name, day"
            ).fetchall()
        con.close()
        return rows
    except Exception as e:
        print(f"  Warning: could not load menus: {e}")
        return []

def _get_weeks():
    """Return sorted list of distinct week_start values in DB."""
    try:
        con = sqlite3.connect(str(DB_PATH))
        weeks = [r[0] for r in con.execute(
            "SELECT DISTINCT week_start FROM client_menus ORDER BY week_start DESC"
        ).fetchall()]
        con.close()
        return weeks
    except Exception:
        return []

# ── Build master workbook ──────────────────────────────────────────────────────
def build_master_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    weeks = _get_weeks()
    clients = _load_clients()

    # Always include current week even if no data yet
    this_week = _monday(date.today()).isoformat()
    if this_week not in weeks:
        weeks.insert(0, this_week)

    for week_start in weeks:
        _build_week_sheet(wb, week_start, clients)

    _build_clients_sheet(wb, clients)

    wb.save(str(MASTER_FILE))
    print(f"  ✅ Saved: {MASTER_FILE}")
    return str(MASTER_FILE)

def _build_week_sheet(wb, week_start: str, clients: dict):
    """Build one week tab — rows = clients, columns = day × course."""
    label = f"Week {week_start}"
    ws = wb.create_sheet(title=label)
    ws.sheet_properties.tabColor = GOLD

    # ── Header row 1: title ───────────────────────────────────────────────────
    monday = datetime.fromisoformat(week_start)
    saturday = monday + timedelta(days=5)
    title = (f"GOJ Menu — Week of {monday.strftime('%B %d')} "
             f"– {saturday.strftime('%B %d, %Y')}")
    ws.merge_cells("A1:AZ1")
    ws["A1"] = title
    _style(ws, "A1", bold=True, size=14, color=WHITE, bg=NAVY, align="center")
    ws.row_dimensions[1].height = 28

    # ── Header row 2: day blocks ──────────────────────────────────────────────
    COL_START = 3  # column C onwards (col A=name, col B=shift)
    ws["A2"] = "Client Name"
    ws["B2"] = "Shift"
    _style(ws, "A2", bold=True, bg=NAVY, color=WHITE, border=True, align="center")
    _style(ws, "B2", bold=True, bg=NAVY, color=WHITE, border=True, align="center")

    col = COL_START
    day_col_map = {}  # day → starting column
    for day in DAYS_ORDER:
        day_date = monday + timedelta(days=DAYS_ORDER.index(day))
        day_label = f"{DAY_LABELS[day]}\n{day_date.strftime('%m/%d')}"
        start_col = col
        end_col   = col + len(COURSES) - 1
        ws.merge_cells(start_row=2, start_column=start_col,
                        end_row=2, end_column=end_col)
        cell = ws.cell(row=2, column=start_col, value=day_label)
        _style(ws, cell, bold=True, bg=NAVY, color=WHITE, align="center",
               border=True, wrap=True)
        day_col_map[day] = start_col
        col = end_col + 1

    ws.row_dimensions[2].height = 36

    # ── Header row 3: course sub-headers ─────────────────────────────────────
    for day in DAYS_ORDER:
        start_col = day_col_map[day]
        for i, course in enumerate(COURSES):
            cell = ws.cell(row=3, column=start_col + i,
                           value=COURSE_LABELS[course])
            _style(ws, cell, bold=True, size=9, bg=LIGHT_NAVY, align="center",
                   border=True, wrap=True)
    ws.row_dimensions[3].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    menus = _load_menus(week_start)
    # Build lookup: client_name → day → course → value
    lookup = defaultdict(lambda: defaultdict(dict))
    for (cname, _, day, salad, soup, main, side, conf) in menus:
        lookup[cname][day] = {
            "salad": salad or "", "soup": soup or "",
            "main": main or "", "side": side or ""
        }

    # Merge clients from DB + clients found in menus this week
    all_names = sorted(set(clients.keys()) | set(lookup.keys()))

    row = 4
    for i, name in enumerate(all_names):
        bg = GRAY_LIGHT if i % 2 == 0 else WHITE
        shift = clients.get(name, {}).get("shift", "?")

        ws.cell(row=row, column=1, value=name)
        _style(ws, ws.cell(row=row, column=1), bold=True, bg=bg, border=True)
        ws.cell(row=row, column=2, value=f"Shift {shift}")
        _style(ws, ws.cell(row=row, column=2), bg=bg, border=True, align="center")

        for day in DAYS_ORDER:
            start_col = day_col_map[day]
            day_data  = lookup[name].get(day, {})
            for j, course in enumerate(COURSES):
                val  = day_data.get(course, "")
                cell = ws.cell(row=row, column=start_col + j, value=val)
                cell_bg = LIGHT_GOLD if val else bg
                _style(ws, cell, bg=cell_bg, border=True, wrap=True, size=9)

        row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    for day in DAYS_ORDER:
        start_col = day_col_map[day]
        for i in range(len(COURSES)):
            ws.column_dimensions[get_column_letter(start_col + i)].width = 18

    ws.freeze_panes = "C4"

def _build_clients_sheet(wb, clients: dict):
    ws = wb.create_sheet(title="Clients")
    ws.sheet_properties.tabColor = NAVY

    headers = ["Client Name", "Shift", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Phone"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _style(ws, cell, bold=True, bg=NAVY, color=WHITE, border=True, align="center")

    ws.row_dimensions[1].height = 22
    day_codes = ["M","T","W","TH","F","SA"]

    for i, (name, info) in enumerate(sorted(clients.items()), 2):
        bg = GRAY_LIGHT if i % 2 == 0 else WHITE
        ws.cell(row=i, column=1, value=name)
        _style(ws, ws.cell(row=i, column=1), bg=bg, border=True)
        ws.cell(row=i, column=2, value=f"Shift {info['shift']}")
        _style(ws, ws.cell(row=i, column=2), bg=bg, border=True, align="center")
        for j, code in enumerate(day_codes):
            val = "✓" if code in info.get("days", set()) else ""
            cell = ws.cell(row=i, column=3+j, value=val)
            cell_bg = GREEN_LIGHT if val else bg
            _style(ws, cell, bg=cell_bg, border=True, align="center")

    ws.column_dimensions["A"].width = 24
    for col in range(2, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.freeze_panes = "A2"

# ── Show week summary in terminal ──────────────────────────────────────────────
def show_week(week_start: str):
    menus = _load_menus(week_start)
    if not menus:
        print(f"No menu data for week of {week_start}")
        return
    print(f"\nMenu week of {week_start}:")
    print(f"{'Client':<25} {'Day':<4} {'Salad':<25} {'Soup':<25} {'Main':<30} {'Side':<20}")
    print("-"*130)
    for r in menus:
        name, _, day, salad, soup, main, side, conf = r
        print(f"{name:<25} {day:<4} {(salad or ''):<25} {(soup or ''):<25} {(main or ''):<30} {(side or ''):<20}")

# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="GOJ Master Menu sync")
    parser.add_argument("--week",  action="store_true", help="Sync current week only")
    parser.add_argument("--show",  action="store_true", help="Print current week to terminal")
    args = parser.parse_args()

    this_week = _monday(date.today()).isoformat()

    if args.show:
        show_week(this_week)
        return

    print(f"  Syncing menu data → {MASTER_FILE.name} ...")
    DASHBOARD.mkdir(parents=True, exist_ok=True)
    path = build_master_workbook()
    print(f"  Done. Open: {path}")

if __name__ == "__main__":
    main()
