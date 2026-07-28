#!/usr/bin/env python3
"""
CC_kanban_drive_monitor.py — Google Drive polling daemon for Gold Health Systems / Garden of Joy.

Polls GOJ Drive folders every 5 minutes for new or changed files.
Always works on a CLONE — the employee owns the originals, this script never touches them.

Watched folders (akhiger@gmail.com, shared with atigerclawai@gmail.com):
  Calendar/Attendance : 1VcNscnjp-rVfUHDxty1g-Njla34uUTTl
  Menus               : 1OBrFP9NR_1lYm_PLHjXXgnISqtxMxuo4
  Sign-In Sheets      : 1znUHkOMfuSQo9iK1Nnz-SSoWVZdnax6H
  Distribution Sheets : 1m8GAglqzBKEdrDuU5Am08Hl9MHnOqhsG
  Kitchen Counts      : 1o56SCqK7QZVcDorAo1oyOAwiEu4CyVu8
  Driver Sheets       : 1JCh5oQt9yJODyLB5PGdTLCjxku3a17HG

Behaviour:
  XLSX in Attendance folder → clone to /tmp/ghs_drive_clone/ → parse → sync clients to auth_tracker.db
  PDF in Menus folder       → clone to ~/Desktop/REX/menus/  → OCR queue picks it up

Rules enforced:
  - Larry NEVER appears on any transport/driver list (blocked at DB write layer)
  - All new files this script creates get CC_ prefix
  - Logs to ~/Desktop/REX/logs/CC_kanban_drive_monitor.log

Usage:
  python3 CC_kanban_drive_monitor.py           # daemon (infinite loop, 5 min poll)
  python3 CC_kanban_drive_monitor.py --once    # single poll then exit
"""

import argparse
import io
import json
import logging
import os
import shutil
import sqlite3
import sys
import time
from datetime import datetime
from pathlib import Path

# ── Paths ──────────────────────────────────────────────────────────────────────

REX_DIR    = Path(__file__).resolve().parent
LOG_DIR    = REX_DIR / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_PATH   = LOG_DIR / "CC_kanban_drive_monitor.log"

CLONE_DIR  = Path("/tmp/ghs_drive_clone")
CLONE_DIR.mkdir(parents=True, exist_ok=True)

MENUS_DIR  = REX_DIR / "menus"
MENUS_DIR.mkdir(parents=True, exist_ok=True)

STATE_FILE = REX_DIR / "CC_drive_monitor_state.json"
AUTH_DB    = Path.home() / "Documents" / "goj files" / "dashboard" / "auth_tracker.db"
TOKEN_PATH = Path.home() / ".rex_google_token.json"
CREDS_PATH = REX_DIR / "google_credentials.json"

POLL_INTERVAL = 300  # 5 minutes

TELEGRAM_CHAT_ID = "5587703834"

# ── Watched folders ────────────────────────────────────────────────────────────

FOLDERS = {
    "attendance":    {"id": "1VcNscnjp-rVfUHDxty1g-Njla34uUTTl", "label": "Calendar/Attendance"},
    "menus":         {"id": "1OBrFP9NR_1lYm_PLHjXXgnISqtxMxuo4", "label": "Menus"},
    "signin":        {"id": "1znUHkOMfuSQo9iK1Nnz-SSoWVZdnax6H", "label": "Sign-In Sheets"},
    "distribution":  {"id": "1m8GAglqzBKEdrDuU5Am08Hl9MHnOqhsG", "label": "Distribution Sheets"},
    "kitchen":       {"id": "1o56SCqK7QZVcDorAo1oyOAwiEu4CyVu8", "label": "Kitchen Counts"},
    "driver":        {"id": "1JCh5oQt9yJODyLB5PGdTLCjxku3a17HG", "label": "Driver Sheets"},
}

# ── Logging ────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("cc-drive-monitor")

# ── State ──────────────────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text())
        except Exception:
            pass
    return {}


def save_state(state: dict) -> None:
    STATE_FILE.write_text(json.dumps(state, indent=2))

# ── Telegram ───────────────────────────────────────────────────────────────────

def notify(msg: str) -> None:
    """Best-effort Telegram notification. Never raises."""
    try:
        import urllib.request
        token = os.environ.get("TELEGRAM_BOT_TOKEN", "")
        if not token:
            log.debug("TELEGRAM_BOT_TOKEN not set — skipping notification")
            return
        payload = json.dumps({"chat_id": TELEGRAM_CHAT_ID, "text": msg}).encode()
        req = urllib.request.Request(
            f"https://api.telegram.org/bot{token}/sendMessage",
            data=payload,
            headers={"Content-Type": "application/json"},
        )
        urllib.request.urlopen(req, timeout=10)
    except Exception as e:
        log.debug(f"Telegram notify failed (non-fatal): {e}")

# ── Google Drive auth ──────────────────────────────────────────────────────────

def _get_drive_service():
    """Return authenticated Drive API service. Refreshes token if expired."""
    try:
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
    except ImportError:
        raise RuntimeError(
            "Google API packages missing. Run:\n"
            "pip install google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client"
        )

    SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
    creds = None

    if TOKEN_PATH.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            TOKEN_PATH.write_text(creds.to_json())
        else:
            raise RuntimeError(
                f"Drive token missing or invalid. Re-auth:\n"
                f"  python backend/rex_gmail.py --setup\n"
                f"Token path: {TOKEN_PATH}"
            )

    return build("drive", "v3", credentials=creds)


# ── Drive helpers ──────────────────────────────────────────────────────────────

def list_folder_files(service, folder_id: str) -> list[dict]:
    """Return all files in a Drive folder (name, id, modifiedTime, mimeType)."""
    results = []
    page_token = None
    while True:
        params = {
            "q": f"'{folder_id}' in parents and trashed=false",
            "fields": "nextPageToken, files(id, name, modifiedTime, mimeType)",
            "pageSize": 200,
        }
        if page_token:
            params["pageToken"] = page_token
        resp = service.files().list(**params).execute()
        results.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return results


def download_file_clone(service, file_id: str, file_name: str, dest_dir: Path) -> Path:
    """
    Download a Drive file into dest_dir with CC_ prefix.
    For Google Sheets → export as XLSX.
    For Google Docs  → export as PDF.
    Never modifies the original.
    Returns the local clone path.
    """
    from googleapiclient.http import MediaIoBaseDownload

    SHEET_MIME = "application/vnd.google-apps.spreadsheet"
    DOC_MIME   = "application/vnd.google-apps.document"

    # Build safe local filename with CC_ prefix
    safe_name = "CC_" + "".join(c if c.isalnum() or c in "._-" else "_" for c in file_name)

    file_meta = service.files().get(fileId=file_id, fields="mimeType").execute()
    mime = file_meta.get("mimeType", "")

    if mime == SHEET_MIME:
        safe_name = safe_name.rstrip("_") + ".xlsx" if not safe_name.endswith(".xlsx") else safe_name
        request = service.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    elif mime == DOC_MIME:
        safe_name = safe_name.rstrip("_") + ".pdf" if not safe_name.endswith(".pdf") else safe_name
        request = service.files().export_media(fileId=file_id, mimeType="application/pdf")
    else:
        request = service.files().get_media(fileId=file_id)

    dest_path = dest_dir / safe_name
    with io.FileIO(dest_path, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()

    log.info(f"  Cloned → {dest_path}")
    return dest_path


# ── Attendance XLSX sync ───────────────────────────────────────────────────────

LARRY_BLOCK = frozenset(["larry"])  # case-insensitive check — NEVER on transport lists


def _larry_blocked(name: str) -> bool:
    return any(part.lower() in LARRY_BLOCK for part in name.split())


def sync_attendance_xlsx(xlsx_path: Path) -> dict:
    """
    Parse attendance XLSX (tabs = day/shift like 'TH 2') and upsert clients into
    auth_tracker.db. Drive is authoritative — DB syncs FROM it.

    Returns {"upserted": int, "skipped": int, "larry_blocked": int}
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl missing. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    upserted = skipped = larry_blocked = 0

    con = sqlite3.connect(str(AUTH_DB))
    cur = con.cursor()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row — look for a row containing "name" (case-insensitive)
        header_idx = None
        header_map = {}
        for i, row in enumerate(rows[:10]):
            row_lower = [str(c).lower() if c else "" for c in row]
            if any("name" in cell for cell in row_lower):
                header_idx = i
                for j, cell in enumerate(row_lower):
                    header_map[cell] = j
                break

        if header_idx is None:
            log.debug(f"  Sheet '{sheet_name}': no header row found — skipping")
            continue

        name_col = next((v for k, v in header_map.items() if "name" in k), None)
        if name_col is None:
            continue

        for row in rows[header_idx + 1:]:
            if not row or len(row) <= name_col:
                continue
            raw_name = row[name_col]
            if not raw_name:
                continue
            full_name = str(raw_name).strip()
            if not full_name:
                continue

            # LARRY RULE — hard block
            if _larry_blocked(full_name):
                larry_blocked += 1
                log.warning(f"  LARRY RULE: blocked '{full_name}' from transport/driver list (sheet: {sheet_name})")
                continue

            # Split into first/last if possible (Drive format: "Last First" or "First Last")
            parts = full_name.split()
            last_name  = parts[0] if parts else full_name
            first_name = " ".join(parts[1:]) if len(parts) > 1 else ""

            try:
                cur.execute(
                    """
                    INSERT INTO clients (first_name, last_name, full_name, status, source, updated_at)
                    VALUES (?, ?, ?, 'ACTIVE', 'drive_sync', ?)
                    ON CONFLICT(full_name) DO UPDATE SET
                        first_name = excluded.first_name,
                        last_name  = excluded.last_name,
                        status     = 'ACTIVE',
                        source     = 'drive_sync',
                        updated_at = excluded.updated_at
                    """,
                    (first_name, last_name, full_name, datetime.utcnow().isoformat()),
                )
                upserted += 1
            except sqlite3.OperationalError as e:
                # Column mismatch — try minimal upsert
                try:
                    cur.execute(
                        """
                        INSERT INTO clients (first_name, last_name)
                        VALUES (?, ?)
                        ON CONFLICT DO NOTHING
                        """,
                        (first_name, last_name),
                    )
                    upserted += 1
                except Exception as e2:
                    log.warning(f"  DB insert failed for '{full_name}': {e2}")
                    skipped += 1
            except Exception as e:
                log.warning(f"  DB insert failed for '{full_name}': {e}")
                skipped += 1

    con.commit()
    con.close()
    wb.close()
    return {"upserted": upserted, "skipped": skipped, "larry_blocked": larry_blocked}


# ── PDF queue drop ─────────────────────────────────────────────────────────────

def queue_pdf_for_ocr(pdf_path: Path) -> None:
    """Drop PDF into menus/ dir. CC_ocr_queue.py picks it up on next run."""
    dest = MENUS_DIR / pdf_path.name
    if not dest.exists():
        shutil.copy2(pdf_path, dest)
        log.info(f"  Queued PDF for OCR: {dest}")

        # Optionally enqueue via CC_ocr_queue if available
        try:
            sys.path.insert(0, str(REX_DIR))
            from CC_ocr_queue import enqueue_scan
            job_id = enqueue_scan(str(dest), "hybrid")
            if job_id:
                log.info(f"  Enqueued as OCR job {job_id}")
        except ImportError:
            log.debug("  CC_ocr_queue not available — PDF dropped in menus/ for manual pickup")
        except Exception as e:
            log.warning(f"  OCR enqueue failed (non-fatal): {e}")
    else:
        log.debug(f"  PDF already present in queue: {dest.name}")


# ── Main poll ──────────────────────────────────────────────────────────────────

def poll(service, state: dict) -> dict:
    """
    Poll all watched folders. Process new/changed files. Return updated state.
    State shape: { folder_key: { file_id: modifiedTime } }
    """
    changed_summaries = []

    for folder_key, folder_info in FOLDERS.items():
        folder_id    = folder_info["id"]
        folder_label = folder_info["label"]
        seen         = state.get(folder_key, {})
        new_seen     = dict(seen)

        try:
            files = list_folder_files(service, folder_id)
        except Exception as e:
            log.error(f"  Failed to list '{folder_label}': {e}")
            continue

        for f in files:
            fid   = f["id"]
            fname = f["name"]
            mtime = f.get("modifiedTime", "")

            if seen.get(fid) == mtime:
                continue  # no change

            is_new = fid not in seen
            action = "NEW" if is_new else "CHANGED"
            log.info(f"[{folder_label}] {action}: {fname}")
            new_seen[fid] = mtime

            # Clone to working dir
            try:
                clone_path = download_file_clone(service, fid, fname, CLONE_DIR)
            except Exception as e:
                log.error(f"  Clone failed for '{fname}': {e}")
                continue

            # Process based on folder type
            try:
                if folder_key == "attendance":
                    ext = clone_path.suffix.lower()
                    if ext in (".xlsx", ".xls", ".csv"):
                        result = sync_attendance_xlsx(clone_path)
                        summary = (
                            f"{action} attendance file: {fname}\n"
                            f"Clients upserted: {result['upserted']} | "
                            f"Skipped: {result['skipped']} | "
                            f"Larry blocked: {result['larry_blocked']}"
                        )
                        log.info(f"  Attendance sync: {result}")
                        changed_summaries.append(summary)
                    else:
                        log.info(f"  Attendance file not XLSX — skipped processing: {fname}")

                elif folder_key == "menus":
                    ext = clone_path.suffix.lower()
                    if ext == ".pdf":
                        queue_pdf_for_ocr(clone_path)
                        changed_summaries.append(f"{action} menu PDF queued for OCR: {fname}")
                    else:
                        log.info(f"  Menu file not PDF — skipped processing: {fname}")

                else:
                    # Other folders (signin, distribution, kitchen, driver): log only
                    summary = f"{action} file in {folder_label}: {fname}"
                    log.info(f"  {summary}")
                    changed_summaries.append(summary)

            except Exception as e:
                log.error(f"  Processing failed for '{fname}': {e}")
                # Don't update state for this file — will retry next poll
                del new_seen[fid]
                continue

        state[folder_key] = new_seen

    if changed_summaries:
        msg = "GHS Drive Monitor\n" + "\n".join(f"• {s}" for s in changed_summaries)
        notify(msg)

    return state


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="GOJ Google Drive polling daemon")
    parser.add_argument(
        "--once", action="store_true",
        help="Run a single poll then exit (default: daemon loop every 5 min)"
    )
    args = parser.parse_args()

    log.info("CC_kanban_drive_monitor starting")
    log.info(f"Clone dir : {CLONE_DIR}")
    log.info(f"State file: {STATE_FILE}")
    log.info(f"Auth DB   : {AUTH_DB}")
    log.info(f"Poll mode : {'once' if args.once else f'every {POLL_INTERVAL}s'}")

    try:
        service = _get_drive_service()
        log.info("Drive auth OK")
    except Exception as e:
        log.error(f"Drive auth failed: {e}")
        sys.exit(1)

    state = load_state()

    if args.once:
        state = poll(service, state)
        save_state(state)
        log.info("Single poll complete. Exiting.")
        return

    # Daemon loop
    notify("GHS Drive Monitor started — polling every 5 min")
    while True:
        try:
            state = poll(service, state)
            save_state(state)
        except KeyboardInterrupt:
            log.info("Interrupted — shutting down")
            notify("GHS Drive Monitor stopped (KeyboardInterrupt)")
            break
        except Exception as e:
            log.error(f"Poll loop error (will retry next cycle): {e}")

        log.info(f"Sleeping {POLL_INTERVAL}s until next poll...")
        try:
            time.sleep(POLL_INTERVAL)
        except KeyboardInterrupt:
            log.info("Interrupted during sleep — shutting down")
            notify("GHS Drive Monitor stopped")
            break


if __name__ == "__main__":
    main()
