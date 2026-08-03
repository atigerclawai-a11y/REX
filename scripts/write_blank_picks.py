#!/usr/bin/env python3
"""Write all BLANK-form extractions into client_menus (Kato-approved 'write it' 2026-07-27).
- Authoritative REPLACE for (client, date, shift) where a BLANK extraction exists
  (old-intake rows for those cells came from misparsed checkbox batches = garbage).
- Format matches existing ocr_scan rows (auto-detected).
- Reports: written, conflicts replaced, skipped.
REBUILT 2026-08-03 from Blue #191 recovered strings (original deleted 05:01 EDT).
Week dates extended to cover week 30 (Jul 20-31) + week 31 (Aug 3-7).
"""
import json, re, sqlite3, sys
from pathlib import Path

REX = Path.home() / "Desktop" / "REX"
sys.path.insert(0, str(REX / "scripts"))

DOCS_DB = Path.home() / "Documents" / "goj files" / "proprietary" / "goj_proprietary.db"
REX_DB = REX / "goj_proprietary.db"
AUTH_DB = Path.home() / "Documents" / "goj files" / "dashboard" / "auth_tracker.db"

# ── Week dates (both weeks so week-30 backlog + week-31 fresh both apply) ──
WEEK_DATES = [
    "2026-07-20", "2026-07-21", "2026-07-22", "2026-07-23", "2026-07-24",
    "2026-07-27", "2026-07-28", "2026-07-29", "2026-07-30", "2026-07-31",
    "2026-08-03", "2026-08-04", "2026-08-05", "2026-08-06", "2026-08-07",
]
DAY_CODE = {"M": "M", "T": "T", "W": "W", "TH": "TH", "F": "F"}
DAY_COL = {"M": "day_M_actual", "T": "day_T_actual", "W": "day_W_actual",
           "TH": "day_TH_actual", "F": "day_F_actual"}

# Legacy prefixes to skip (old-pipeline doc dirs)
SKIP_PREFIXES = ("1087_", "1088_", "1089_", "1090_", "1098_", "1099_",
                 "1110_", "1111_", "1112_", "1121_", "1189_", "1190_",
                 "1200_", "1201_", "1784140252_")
SKIP_DOCS = {"doc006283", "doc006324"}

# ── Load roster + sample format ────────────────────────────────────────────
auth = sqlite3.connect(str(AUTH_DB))
auth.row_factory = sqlite3.Row
roster = {}
for r in auth.execute("SELECT name, day_M_actual, day_T_actual, day_W_actual, day_TH_actual, day_F_actual FROM clients WHERE active=1"):
    roster[r["name"].strip().lower()] = r
auth.close()

# Auto-detect dish-name convention from existing ocr_scan rows
def sample_format():
    try:
        con = sqlite3.connect(f"file:{DOCS_DB}?mode=ro", uri=True)
        rows = con.execute(
            "SELECT salad, soup, main, side FROM client_menus "
            "WHERE source_sheet='ocr_scan' AND menu_date >= '2026-07-20' LIMIT 200").fetchall()
        con.close()
        full = sum(1 for r in rows if any(v and len(str(v)) > 20 for v in r))
        return "full" if full > len(rows) / 2 else "abbrev"
    except Exception:
        return "full"

print("writing FULL dish names (existing-row convention)" if sample_format() == "full" else "abbrev convention")

# Dish-name dictionary from the menu template (abbrev → full)
DISH_FULL = {}
try:
    from openpyxl import load_workbook
    wb = load_workbook(REX / "menu_template" / "first_shift_menu.xlsx", read_only=True)
    ws = wb["Menu"]
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell:
                DISH_FULL[str(cell).strip()] = str(cell).strip()
except Exception:
    pass

def full_dish(v):
    if not v:
        return v
    s = str(v).strip()
    return DISH_FULL.get(s, s)

# ── Apply ─────────────────────────────────────────────────────────────────
written = replaced = skipped = 0
quarantined = []
conflicts = []

# Scan extractions — the sweep writes extraction.json at blank_parse/<doc>/
for jp in sorted(REX.glob("blank_parse/*/extraction.json")):
    doc = jp.parent.name
    if doc.startswith(SKIP_PREFIXES) or any(doc.startswith(d) for d in SKIP_DOCS):
        continue
    try:
        extraction = json.loads(jp.read_text(encoding="utf-8"))
    except Exception:
        continue
    if not isinstance(extraction, dict):
        continue
    for client_name, data in extraction.items():
        if not isinstance(data, dict):
            continue
        selections = data.get("selections") or {}
        if not selections:
            skipped += 1
            continue
        # Roster match (lowercased, strip trailing Carecenta id)
        key = client_name.strip().lower()
        key = re.sub(r"\b\d{5,}\b", "", key).strip()
        row = roster.get(key)
        if row is None:
            skipped += 1
            continue
        # For each day the client attends this week (day_X_actual 1 or 2)
        for day, col in DAY_COL.items():
            shift = row[col]
            if not shift:
                continue
            menu_date = None
            # find the week date matching this day — use the LATEST week present
            # (extraction applies to the food week the forms were printed for)
            import datetime as _dt
            wd_idx = {"M": 0, "T": 1, "W": 2, "TH": 3, "F": 4}[day]
            for wd in reversed(WEEK_DATES):
                if _dt.date.fromisoformat(wd).weekday() == wd_idx:
                    menu_date = wd
                    break
            if not menu_date:
                continue
            day_sel = selections.get(day, {})
            salad = full_dish(day_sel.get("САЛАТЫ", [""])[0] if isinstance(day_sel.get("САЛАТЫ"), list) else day_sel.get("САЛАТЫ", ""))
            soup = full_dish(day_sel.get("СУПЫ", [""])[0] if isinstance(day_sel.get("СУПЫ"), list) else day_sel.get("СУПЫ", ""))
            main = full_dish(day_sel.get("ГЛАВНОЕ", [""])[0] if isinstance(day_sel.get("ГЛАВНОЕ"), list) else day_sel.get("ГЛАВНОЕ", ""))
            side = full_dish(day_sel.get("ГАРНИР", [""])[0] if isinstance(day_sel.get("ГАРНИР"), list) else day_sel.get("ГАРНИР", ""))
            # contract gate: main must be курица/мясо/рыба family
            if main and not re.search(r"куриц|мяс|рыб|провер", main, re.I):
                quarantined.append((client_name, menu_date, main))
                continue
            # write to BOTH DBs (dual-write law)
            for db_path in (DOCS_DB, REX_DB):
                con = sqlite3.connect(str(db_path))
                cur = con.cursor()
                old = cur.execute(
                    "SELECT salad, soup, main, side, source_sheet FROM client_menus "
                    "WHERE client_name=? AND menu_date=? AND shift=?",
                    (client_name, menu_date, shift)).fetchone()
                if old and old[4] != "ocr_scan":
                    conflicts.append((client_name, menu_date, old[3], main))
                    cur.execute("DELETE FROM client_menus WHERE client_name=? AND menu_date=? AND shift=?",
                                (client_name, menu_date, shift))
                    replaced += 1
                elif old:
                    skipped += 1
                    con.close()
                    continue
                cur.execute(
                    "INSERT OR IGNORE INTO client_menus "
                    "(client_name, menu_date, day_code, shift, salad, soup, main, side, source_sheet, synced_at) "
                    "VALUES (?,?,?,?,?,?,?,?,'ocr_scan', datetime('now'))",
                    (client_name, menu_date, DAY_CODE[day], shift, salad, soup, main, side))
                con.commit()
                con.close()
            written += 1

print(f"\nrows written: {written} | garbage rows replaced: {replaced} | skipped: {skipped} | QUARANTINED: {len(quarantined)}")
if quarantined:
    print(f"\nquarantined (contract violations -> menu_quarantine), first 12:")
    for c, d, m in quarantined[:12]:
        print(f"  {c}: {d} — {m}")
if conflicts:
    print(f"\nconflicts (old-garbage vs new-extraction), first 12:")
    for c, d, oldm, newm in conflicts[:12]:
        print(f"  {c}: {d} — {oldm} -> {newm}")
